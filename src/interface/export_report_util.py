"""导出报告工具 - Export Report Utility
支持 Excel (.xlsx) 和 PDF (.pdf) 格式导出
"""

import os
import tempfile
from io import BytesIO
from typing import List, Dict, Any

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei"]
plt.rcParams["axes.unicode_minus"] = False

from .styles import COLORS


def _generate_focus_chart_image(records: list) -> BytesIO:
    """生成专注度趋势图（所有曲线），返回 PNG 字节流"""
    fig = plt.Figure(figsize=(10, 5), facecolor="white")
    ax = fig.add_subplot(111)
    ax.set_facecolor("white")

    cc = COLORS["chart_colors"]

    sampled = _downsample(records, 60)
    n = len(sampled)
    if n == 0:
        ax.text(0.5, 0.5, "No data", ha="center", va="center", fontsize=14)
        buf = BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
        buf.seek(0)
        plt.close(fig)
        return buf

    x = np.arange(n)

    line_configs = [
        ("final_focus_score", "最终专注度", cc[0], 2.5),
        ("head_pose_score", "头部姿态", cc[1], 1.5),
        ("behavior_score", "行为动作", cc[2], 1.5),
        ("expression_score", "表情评分", cc[3], 1.5),
        ("evidence_score", "证据理论", cc[4], 1.5),
        ("people_score", "人数项", cc[5], 1.5),
    ]

    for key, label, color, lw in line_configs:
        ax.plot(x, [r.get(key, 0) for r in sampled],
                label=label, color=color, linewidth=lw,
                linestyle="-" if key == "final_focus_score" else "--")

    ax.set_title("专注度评分变化趋势", fontsize=14, color="#333333")
    ax.set_xlabel("采样点", fontsize=10, color="#666666")
    ax.set_ylabel("评分", fontsize=10, color="#666666")
    ax.set_ylim(0, 100)
    ax.grid(True, linestyle="--", alpha=0.3)
    ax.tick_params(axis="both", colors="#666666")
    for spine in ax.spines.values():
        spine.set_color("#CCCCCC")

    ax.legend(loc="upper right", fontsize=9,
              facecolor="white", edgecolor="#CCCCCC")

    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def _downsample(records: list, max_samples: int) -> list:
    n = len(records)
    if n <= max_samples:
        return records
    step = n // max_samples
    return records[::step][:max_samples]


def export_to_excel(session: dict, records: list, alerts: list, filepath: str):
    """导出 Excel 报告（三个 Sheet）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()

    # ── 通用样式 ──
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="7A5CFF", end_color="7A5CFF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def _write_header(ws, row, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border

    def _write_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = cell_font
        c.alignment = cell_align
        c.border = thin_border
        return c

    # ═══════════════════════════════════════
    # Sheet 1: 课程概述
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "课程概述"

    info_headers = ["字段", "内容"]
    _write_header(ws1, 1, info_headers)
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 36

    session_id = session.get("session_id", "")
    start_time = session.get("start_time", "")
    end_time = session.get("end_time", "")
    mode_map = {"class": "网课模式", "exam": "考试模式"}
    mode = mode_map.get(session.get("mode", ""), session.get("mode", ""))
    avg_focus = session.get("avg_focus_score", 0)
    abnormal_count = session.get("abnormal_event_count", 0)
    date_str = start_time.split(" ")[0] if " " in start_time else start_time
    time_start = start_time.split(" ")[1] if " " in start_time else start_time
    time_end = end_time.split(" ")[1] if " " in end_time else end_time

    fields = [
        ("会话 ID", session_id),
        ("日期", date_str),
        ("开始时间", time_start),
        ("结束时间", time_end),
        ("模式", mode),
        ("平均专注度", f"{avg_focus:.1f}"),
        ("异常事件数", str(abnormal_count)),
        ("专注度采样点数", str(len(records))),
    ]

    for i, (field, value) in enumerate(fields, 2):
        _write_cell(ws1, i, 1, field)
        _write_cell(ws1, i, 2, value)

    # ═══════════════════════════════════════
    # Sheet 2: 关键帧专注度
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("关键帧专注度")

    record_headers = [
        "时间戳(s)", "头部姿态", "行为动作", "表情",
        "证据理论", "人数项", "最终专注度", "强制置0"
    ]
    col_widths = [14, 12, 12, 12, 12, 12, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    _write_header(ws2, 1, record_headers)

    for row, record in enumerate(records, 2):
        values = [
            f"{record.get('timestamp', 0):.1f}",
            f"{record.get('head_pose_score', 0):.1f}",
            f"{record.get('behavior_score', 0):.1f}",
            f"{record.get('expression_score', 0):.1f}",
            f"{record.get('evidence_score', 0):.1f}",
            f"{record.get('people_score', 0):.1f}",
            f"{record.get('final_focus_score', 0):.1f}",
            "是" if record.get("is_force_zero", False) else "否",
        ]
        for col, val in enumerate(values, 1):
            _write_cell(ws2, row, col, val)

    # 图表插入
    chart_buf = _generate_focus_chart_image(records)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(chart_buf.read())
        tmp_path = tmp.name

    img = XLImage(tmp_path)
    img.width = 600
    img.height = 300
    chart_start_row = len(records) + 4
    ws2.add_image(img, f"A{chart_start_row}")

    # ═══════════════════════════════════════
    # Sheet 3: 告警记录
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("告警记录")

    alert_headers = ["课程ID", "时间", "类型", "详细描述"]
    alert_widths = [18, 14, 18, 36]
    for i, w in enumerate(alert_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    _write_header(ws3, 1, alert_headers)

    if alerts:
        for row, alert in enumerate(alerts, 2):
            sid = alert.get("session_id", "")
            short_sid = sid[-8:] if len(sid) >= 8 else sid
            ts = alert.get("timestamp", 0)
            time_str = f"{int(ts // 60):02d}:{int(ts % 60):02d}" if ts else ""
            values = [
                short_sid,
                time_str,
                alert.get("alert_type", ""),
                alert.get("detail", ""),
            ]
            for col, val in enumerate(values, 1):
                _write_cell(ws3, row, col, val)
    else:
        _write_cell(ws3, 2, 1, "暂无告警记录")

    try:
        wb.save(filepath)
    finally:
        os.unlink(tmp_path)


def export_to_pdf(session: dict, records: list, alerts: list, filepath: str):
    """导出 PDF 报告（三个部分）"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.colors import HexColor, black, white
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, PageBreak,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus.flowables import KeepTogether

    # ── 注册中文字体 ──
    _register_chinese_font()

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ChTitle", parent=styles["Title"],
        fontName="MicrosoftYaHei", fontSize=18, textColor=HexColor("#333333"),
        spaceAfter=6 * mm,
    )
    section_style = ParagraphStyle(
        "ChSection", parent=styles["Heading2"],
        fontName="MicrosoftYaHei", fontSize=14, textColor=HexColor("#7A5CFF"),
        spaceBefore=5 * mm, spaceAfter=3 * mm,
    )
    body_style = ParagraphStyle(
        "ChBody", parent=styles["Normal"],
        fontName="MicrosoftYaHei", fontSize=10, leading=16,
        textColor=HexColor("#333333"),
    )
    center_style = ParagraphStyle(
        "ChCenter", parent=body_style,
        alignment=TA_CENTER,
    )

    session_id = session.get("session_id", "")
    start_time = session.get("start_time", "")
    end_time = session.get("end_time", "")
    mode_map = {"class": "网课模式", "exam": "考试模式"}
    mode = mode_map.get(session.get("mode", ""), session.get("mode", ""))
    avg_focus = session.get("avg_focus_score", 0)
    abnormal_count = session.get("abnormal_event_count", 0)

    elements = []

    # ── 封面标题 ──
    elements.append(Paragraph("网课专注度分析报告", title_style))
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(f"会话: {session_id}", body_style))
    elements.append(Spacer(1, 6 * mm))

    # ═══════════════════════════════════════
    # Part 1: 课程概述
    # ═══════════════════════════════════════
    elements.append(Paragraph("一、课程概述", section_style))
    elements.append(Spacer(1, 2 * mm))

    overview_data = [
        ["会话 ID", session_id],
        ["日期", start_time.split(" ")[0] if " " in start_time else start_time],
        ["开始时间", start_time.split(" ")[1] if " " in start_time else start_time],
        ["结束时间", end_time.split(" ")[1] if " " in end_time else end_time],
        ["模式", mode],
        ["平均专注度", f"{avg_focus:.1f}"],
        ["异常事件数", str(abnormal_count)],
        ["专注度采样点数", str(len(records))],
    ]

    overview_table = Table(overview_data, colWidths=[5 * cm, 8 * cm])
    overview_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "MicrosoftYaHei"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), HexColor("#666666")),
        ("TEXTCOLOR", (1, 0), (1, -1), HexColor("#333333")),
        ("BACKGROUND", (0, 0), (-1, -1), white),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#E0E0E0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(overview_table)
    elements.append(Spacer(1, 8 * mm))

    # ═══════════════════════════════════════
    # Part 2: 关键帧专注度
    # ═══════════════════════════════════════
    elements.append(Paragraph("二、关键帧专注度", section_style))
    elements.append(Spacer(1, 2 * mm))

    # 记录表格（取前 30 条摘要避免过长）
    table_headers = ["时间戳", "头部姿态", "行为动作", "表情", "证据理论", "人数项", "最终专注度"]
    table_data = [table_headers]
    display_records = records[:30]
    for r in display_records:
        table_data.append([
            f"{r.get('timestamp', 0):.1f}s",
            f"{r.get('head_pose_score', 0):.1f}",
            f"{r.get('behavior_score', 0):.1f}",
            f"{r.get('expression_score', 0):.1f}",
            f"{r.get('evidence_score', 0):.1f}",
            f"{r.get('people_score', 0):.1f}",
            f"{r.get('final_focus_score', 0):.1f}",
        ])

    avail = doc.width
    col_w = avail / 7.0
    record_table = Table(table_data, colWidths=[col_w] * 7, repeatRows=1)
    record_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "MicrosoftYaHei"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#7A5CFF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("TEXTCOLOR", (0, 1), (-1, -1), HexColor("#333333")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#E0E0E0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(record_table)

    if len(records) > 30:
        elements.append(Paragraph(f"... 共 {len(records)} 条，仅显示前 30 条", center_style))

    elements.append(Spacer(1, 5 * mm))

    # 专注度趋势图
    chart_buf = _generate_focus_chart_image(records)
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp.write(chart_buf.read())
        tmp_path = tmp.name

    img = Image(tmp_path, width=doc.width * 0.9, height=doc.width * 0.45)
    elements.append(img)

    elements.append(Spacer(1, 8 * mm))

    # ═══════════════════════════════════════
    # Part 3: 告警记录
    # ═══════════════════════════════════════
    elements.append(Paragraph("三、告警记录", section_style))
    elements.append(Spacer(1, 2 * mm))

    alert_headers = ["课程ID", "时间", "类型", "详细描述"]
    alert_data = [alert_headers]

    if alerts:
        for alert in alerts:
            sid = alert.get("session_id", "")
            short_sid = sid[-8:] if len(sid) >= 8 else sid
            ts = alert.get("timestamp", 0)
            time_str = f"{int(ts // 60):02d}:{int(ts % 60):02d}" if ts else ""
            alert_data.append([
                short_sid, time_str,
                alert.get("alert_type", ""),
                alert.get("detail", ""),
            ])
    else:
        alert_data.append(["", "", "暂无告警记录", ""])

    alert_col_w = [avail * 0.15, avail * 0.12, avail * 0.18, avail * 0.55]
    alert_table = Table(alert_data, colWidths=alert_col_w, repeatRows=1)
    alert_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "MicrosoftYaHei"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#7A5CFF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("TEXTCOLOR", (0, 1), (-1, -1), HexColor("#333333")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#E0E0E0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(alert_table)

    try:
        doc.build(elements)
    finally:
        os.unlink(tmp_path)


_CHINESE_FONT_REGISTERED = False


def _register_chinese_font():
    """注册中文字体到 reportlab"""
    global _CHINESE_FONT_REGISTERED
    if _CHINESE_FONT_REGISTERED:
        return

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_paths = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/msyh.ttf",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/simsun.ttc",
    ]
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("MicrosoftYaHei", fp))
                _CHINESE_FONT_REGISTERED = True
                return
            except Exception:
                continue

    _CHINESE_FONT_REGISTERED = True
